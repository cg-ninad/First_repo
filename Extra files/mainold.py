try:
    import uuid
    import os
    import glob
    import openpyxl
    import datetime
    import time
    import numpy as np
    import pandas as pd
    from openpyxl import load_workbook
    from logger_format import setup_logging
    from FileOperations import FileOperations
    from MailList import MailList
    from Config import FILEPATH
    from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
    from xlsx2html import xlsx2html
    from Config import BUSINESS, CC, BCC, FILEPATH, ETHICS_PATH, ARCHIEVE
except ImportError as error:
    print(f"Import Error..... Script Stopped.. {error}")
    exit(1)


class Main:

    def __init__(self, logFolder=None, logger=None, guid=None):

        self.start_time = time.time()
        if guid:
            self.guid = guid
        else:
            self.guid = uuid.uuid4()
        if logFolder:
            self.logger = logger
        else:
            self.logger = setup_logging(__file__)
        self.extra_dict_common = {
            "JobId": self.guid,
            "logType": "User",
            "exec_category": "Python_Script",
            "Category": "Self Heal",
            "businessCategory": "GI",
            "useCaseName": "GI-Ethics",
        }
        self.logger.info("-------------------------------------START------------------------", extra=self.extra_dict_common)
        self.file_obj = FileOperations(__file__, self.logger, self.guid)
        self.mailobj = MailList(__file__, self.logger, self.guid)
        self.status_file = open("Status.txt",mode="w+")

    def speakup_process(self):
        """
        Speak Up process starts from here.
        :return:
        """
        ethics_file = ''
        try:
            ethics_file = self.file_obj.get_file_name(ETHICS_PATH)
            if False in ethics_file:
                self.logger.error(f"speakup_process(); File not found: {ethics_file}", extra=self.extra_dict_common)
                resp = self.mailobj.drt_file_missing(ethics_file[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                ethics_file = ethics_file[1]
            self.logger.info(f"speakup_process(); ECO_List file name received; {ethics_file}",
                             extra=self.extra_dict_common)
            response = self.file_obj.get_file_name(FILEPATH)

            if False in response:
                resp = self.mailobj.drt_file_missing(response[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                filename = response[1]
                if 'speakup' not in str(os.path.basename(filename)).lower():
                    msg = "SpeakUp file not found in folder, Invalid file is uploaded. Please upload SpeakUp " \
                          f"alerts file including <mark>'SpeakUp'</mark> keyword in filename.<br> current uploaded file: {filename}"
                    resp = self.mailobj.drt_file_missing(msg)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False
            self.logger.info(f"speakup_process(); ECO_List file name received; {filename}",
                             extra=self.extra_dict_common)
            ip_file = self.add_region_in_df_basedon_code(filename, ethics_file)

            resp_df = self.file_obj.read_file(ip_file)
            if len(resp_df) > 1:
                resp = self.mailobj.drt_file_missing(resp_df[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                resp_df = resp_df[0]

            ip_file = self.filter_data_ownerwise(resp_df)
            if len(ip_file) > 1:
                resp = self.mailobj.drt_file_missing(ip_file[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                ip_file = ip_file[0]

            ip_file = self.split_alerts_by_region(ip_file)
            if len(ip_file) > 1:
                resp = self.mailobj.drt_file_missing(ip_file[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                ip_file = ip_file[0]

            wb = load_workbook(ip_file)
            ws_list = wb.worksheets
            for ws in ws_list:
                self.set_border(ws, wb, ip_file)
                self.excel_font_styles(ip_file)
                flg = self.color_formatting(wb, ws, ip_file)
                if not flg:
                    resp = self.mailobj.file_missing_declare(flg)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False

            ip_file = self.set_header_color(ip_file)
            if len(ip_file) > 1:
                # Mail need to be send with error to business.
                resp = self.mailobj.drt_file_missing(ip_file[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                ip_file = ip_file[0]

            flg = self.delete_region_column(ip_file)
            if not flg:
                resp = self.mailobj.drt_file_missing(flg)
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False


            grp_ethics_cc = pd.read_excel(ethics_file, sheet_name='Ethics Team', usecols=["Email Address"])[
                'Email Address'].to_list()
            eco_df = pd.read_excel(ethics_file, sheet_name='ECO', usecols=["Region", "Email Address"])
            eco_dict_cc = {k: list(set(g["Email Address"].tolist())) for k, g in eco_df.groupby("Region")}
            ind_ethics_ids = pd.read_excel(ethics_file, sheet_name='India Ethics Team', usecols=["Email Address"])[
                'Email Address'].to_list()
            ind_hr_ids = pd.read_excel(ethics_file, sheet_name='India HR Team', usecols=["Email Address"])[
                'Email Address'].to_list()


            state = self.process_speakup(ip_file, eco_dict_cc, grp_ethics_cc, ind_ethics_ids, ind_hr_ids, ethics_file)

            # if not state:
            #     resp = self.mailobj.speakup_file_missing(state)
            #     self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
            #     return False

            self.logger.info("speakup_process(); Processing India Region alerts", extra=self.extra_dict_common)

            check_india_file = load_workbook(ip_file)
            check_india_file_list = check_india_file.sheetnames
            if "India" in check_india_file_list:
                self.logger.info("speakup_process(); India sheet present in the input file", extra=self.extra_dict_common)
                print(True)

                india_file = self.add_india_teams_to_file(filename=ip_file, ethics_file=ethics_file)

                if len(india_file) > 1:
                    resp = self.mailobj.drt_file_missing(india_file)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False
                else:
                    india_file = india_file[0]

                india_file = self.__split_data_by_team__(india_file)
                if len(india_file) > 1:
                    resp = self.mailobj.drt_file_missing(india_file)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False
                else:
                    india_file = india_file[0]

                wb = load_workbook(india_file)
                ws_list = wb.worksheets
                for ws in ws_list:
                    self.set_border(ws, wb, india_file)
                    self.excel_font_styles(india_file)
                    flg = self.color_formatting(wb, ws, india_file)
                    if not flg:
                        resp = self.mailobj.file_missing_declare(flg)
                        self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                        return False

                india_file = self.set_header_color(india_file)
                if len(india_file) > 1:
                    resp = self.mailobj.file_missing_declare(india_file)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False
                else:
                    india_file = india_file[0]
                #exit()

                flg = self.delete_region_column_forindia(india_file)
                if not flg:
                    resp = self.mailobj.file_missing_declare(flg)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False

                flg = self.india_region_process_data(filename=india_file, ecos=eco_dict_cc, ethics_team=grp_ethics_cc,
                                                     ind_ethics=ind_ethics_ids, ind_hr=ind_hr_ids, ethics_file=ethics_file)
                if not flg:
                    resp = self.mailobj.file_missing_declare(flg)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False


                subject, mail_body = self.mailobj.speakup_success()
                self.status_file.close()
                attach = [os.path.basename("Status.txt"),os.path.basename("input_data.xlsx"), os.path.basename("India.xlsx")]
                status = self.mailobj.send_mail(subject, mail_body, BUSINESS, cc=CC, bcc=BCC, attachments= attach)


                if status:
                    resp = self.file_obj.move_files_archive(filename, ARCHIEVE)

                else:
                    return False
                return status

            else:

                subject, mail_body = self.mailobj.speakup_success()
                self.status_file.close()
                attach = [os.path.basename("Status.txt"), os.path.basename("input_data.xlsx")]
                status = self.mailobj.send_mail(subject, mail_body, BUSINESS, cc=CC, bcc=BCC, attachments=attach)
                if status:
                    resp = self.file_obj.move_files_archive(filename, ARCHIEVE)

                else:
                    return False
                return status


        except Exception as err:
                self.logger.exception(f"speakup_process(); Exception Raised: {err}", extra=self.extra_dict_common)
                resp = self.mailobj.drt_file_missing(err)
                # self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False

    def filter_data_ownerwise(self, df):
        """
        Filter input dataframe by name Anne-Violaine AGAZZI-MONNIE and save it in different sheets.
        :param df:
        :return:
        """
        try:
            ip_file = "input_data.xlsx"
            writer = pd.ExcelWriter(ip_file, engine='openpyxl')
            anne = [75465]
            bejoy = [1580102]
            both_together = [1580102, 75465]
            df.loc[df['Issue Owner Employee Id'].isin(anne)].to_excel(writer, 'Anne-Violaine',
                                                                                     index=False)
            df.loc[df['Issue Owner Employee Id'].isin(bejoy)].to_excel(writer, 'Bejoy Das', index=False)
            df.loc[~df['Issue Owner Employee Id'].isin(both_together)].to_excel(writer, 'input',
                                                                                                   index=False)
            writer.save()
            return [os.path.abspath(ip_file)]
        except Exception as err:
            self.logger.exception(f"filter_data(); failed to filter out data; {err}", extra=self.extra_dict_common)
            return [False, err]

    def add_region_in_df_basedon_code(self, ip_file=None, ethics_file=None, **krgs):

        try:


            print(ip_file)
            declare_df = pd.read_excel(ip_file)
            declare_df = declare_df.dropna(subset=['Issue Owner'])
            ethics_df = pd.read_excel(ethics_file, usecols=["Country", "Region"])
            region_dict = dict(ethics_df.values)
            print(region_dict)
            #exit()
            declare_df['Region'] = declare_df['Country'].map(region_dict)
            writer = pd.ExcelWriter(ip_file)
            declare_df.to_excel(writer, 'input', index=False)
            writer.save()
            return ip_file
        except Exception as excpt:
            self.logger.exception(f"add_region_in_df_basedon(); Exception: {excpt}", extra=self.extra_dict_common)
            return [False, excpt.__str__()]

    def set_border(self, ws, wb, filename):



        border = Border(left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))
        text_ft = Font(name='Verdana', size=10)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.font = text_ft
                cell.alignment = Alignment(horizontal='center', vertical='center')
            wb.save(filename)

    def color_formatting(self, wb, ws, filename):
        """
        format excel cell based on rules like.

        Rule 1: Update Alert column text color to RED if created date is more than 30 days.
        :return:
        updated file name to share with business.
        """
        try:
            red_ft = Font(color="00FF0000", name='Verdana', size=10) # , size=12
            yellow_ft = openpyxl.styles.colors.Color(rgb='00FFF000')
            yellow_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=yellow_ft)
            for row in ws.iter_rows(min_row=2):
                try:
                    row = list(row)
                    alert_cell = row[0]
                    date_cell = row[1]
                    #sev_cell = row[5]
                    #dept_cell = row[6]
                    #issue_cell = row[7]
                    intake_cell = row[2]
                    issue_owner = row[3]
                    # print(issue_owner.value, type(issue_owner))
                    # encoded_value = str(issue_owner.value).encode('UTF-8').decode("UTF-8")
                    # issue_owner.value = str(encoded_value)
                    # print(str(issue_owner.value).encode('utf-8'))
                    # print(issue_owner.value)


                    intake_type = str(intake_cell.value).capitalize()
                    #issue_cells = str(issue_cell).capitalize()
                    #print("Intake Type:", intake_type)
                    #print("Issue Cell:", issue_cells)

                    delta = datetime.datetime.today() - date_cell.value
                    if int(delta.days) > 30:
                        alert_cell.font = red_ft

                    if (int(delta.days) > 5) and ('question' in intake_type.lower()):
                        alert_cell.font = red_ft
                    """

                    if 'question' not in intake_type.lower():
                        if sev_cell.value is None:
                            sev_cell.fill = yellow_fill
                        if dept_cell.value is None or 'XXX' in str(dept_cell.value).upper():
                            dept_cell.fill = yellow_fill
                        if issue_cell.value is None:
                            issue_cell.fill = yellow_fill
                    elif 'question' in intake_type.lower():
                        if sev_cell.value is None:
                            sev_cell.value = 'NA'
                        if dept_cell.value is None:
                            dept_cell.value = 'NA'
                        if issue_cell.value is None:
                            issue_cell.value = 'NA'"""
                    #wb.save(filename)
                except Exception as e:
                    self.logger.error(f"color_formatting(); failed to process row: {e}", extra=self.extra_dict_common)
            wb.save(filename)
            return True
        except Exception as err:
            self.logger.exception(f"Failed to process: {err}", extra=self.extra_dict_common)
            return err.__str__()

    def excel_font_styles(self, filename):

        try:
            font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                        strike=False)
            border = Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
            wb = load_workbook(filename)
            ws_list = wb.worksheets
            for ws in ws_list:
                for row in ws.iter_rows(min_row=2):
                    for cel in row:
                        cel.font = font
                        cel.border = border
                wb.save(filename)
            return True
        except Exception as err:
            self.logger.exception(f"excel_font_styles(); Exception: {err}", extra=self.extra_dict_common)
            return False, err

    def split_alerts_by_region(self, filename, **krgs):

        try:
            countrywise_file = "input_data.xlsx"
            df = pd.read_excel(filename, sheet_name="input")
            violaine_df = pd.read_excel(filename, sheet_name="Anne-Violaine")
            bejoy_df = pd.read_excel(filename, sheet_name="Bejoy Das")
            dict_of_country = dict(iter(df.groupby('Region')))
            writer = pd.ExcelWriter(countrywise_file, engine='openpyxl')
            for key, val in dict_of_country.items():
                val.to_excel(writer, sheet_name=key, header=True, index=False)

            violaine_df.to_excel(writer, sheet_name='Anne-Violaine', header=True, index=False)
            bejoy_df.to_excel(writer, sheet_name='Bejoy Das', header=True, index=False)
            writer.save()
            #exit()
            return [os.path.abspath(countrywise_file)]
        except Exception as err:
            self.logger.exception(f"filter_countrywise_data(); Failed to Process Data; {err}",
                                  extra=self.extra_dict_common)
            return [False, err]

    def set_header_color(self, fname):

        try:
            bg_clr = openpyxl.styles.colors.Color(rgb='000066CC')
            head_color = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=bg_clr)

            wb = load_workbook(fname)
            ws_list = wb.worksheets
            for ws in ws_list:
                for row in ws.iter_rows(max_row=1):
                    for cel in row:
                        cel.font = Font(name='Calibri', color="00FFFFFF", size=10)
                        cel.fill = head_color
                wb.save(fname)

            return [fname]
        except Exception as err:
            self.logger.exception(f"set_header_color(); Exception: {err}", extra=self.extra_dict_common)
            return [False, err]

    def delete_region_column(self, filename):

        try:
            book = openpyxl.load_workbook(filename)
            ws_list = book.worksheets
            # delete column from existing sheet
            for ws in ws_list:
                ws.delete_cols(8)
                #ws.delete_cols(11)
                book.save(filename)
            return True
        except Exception as err:
            self.logger.exception(f"Delete Col Failed: {err}", extra=self.extra_dict_common)
            return [False, err]

    def delete_region_column_forindia(self, filename):

        try:
            book = openpyxl.load_workbook(filename)
            ws_list = book.worksheets
            # delete column from existing sheet
            for ws in ws_list:
                ws.delete_cols(8)
                #ws.delete_cols(11)
                book.save(filename)
            return True
        except Exception as err:
            self.logger.exception(f"Delete Col Failed: {err}", extra=self.extra_dict_common)
            return [False, err]

    def process_speakup(self, filename, ecos=None, ethics_team=None, ind_ethics=None, ind_hr=None, ethics_file=None,
                        **krgs):

        try:
            xl = pd.ExcelFile(filename)
            ws_list = xl.sheet_names
            print("WS_LIST",ws_list)

            for ws in ws_list:
                try:
                    issue_owners = []
                    #df = xl.parse(ws)
                    df = pd.read_excel(filename, sheet_name=ws)
                    df['Issue Owner Employee Id']= df['Issue Owner Employee Id'].replace(np.nan, 'Empty')
                    print("Dataframe is..",df)
                    emps = list(dict.fromkeys(df['Issue Owner Employee Id'].tolist()))
                    print("Employeesss:", emps)
                    df.drop(['Issue Owner Employee Id'],axis=1,inplace=True)

                    for emp in emps:
                        if emp == 'Empty':
                            continue
                        print("Employee in employee",emp)
                        emp = float(emp)
                        emp = int(emp)
                        issue_owners.append(self.file_obj.get_email_id(emp))
                        print(issue_owners)

                    if ws.__contains__('Anne-Violaine'):
                        self._delete_emp_id_(filename, ws)
                        val_df = pd.read_excel(filename,sheet_name=ws)
                        if len(val_df) > 0:
                            ethics_team = list(np.setdiff1d(ethics_team, issue_owners))
                            self.logger.info(f"TO: {issue_owners}; cc: {ethics_team} [{ethics_team + issue_owners}]")
                            xlsx2html(filename, sheet=ws, output='table.html')
                            body_table = open(os.path.abspath('table.html')).read()
                            mail_body = self.mailobj.speakup_mail(body_table, 'GROUP')
                            flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=issue_owners, cc=ethics_team, bcc=BCC,
                                                         attachments=['sign.png', 'Picture1.png'])
                            self.status_file.write(str(mail_body[0] + " : " + "Success" + "\n"))
                            if not flg:
                                self.logger.error("process_speakup(); Unable to send mail", extra=self.extra_dict_common)
                                return False
                    elif ws.__contains__('Bejoy Das'):
                        self._delete_emp_id_(filename, ws)
                        val_df = pd.read_excel(filename, sheet_name=ws)
                        if len(val_df) > 0:
                            ethics_team = list(np.setdiff1d(ethics_team, issue_owners))
                            self.logger.info(f"TO: {issue_owners}; cc: {ethics_team} [{ethics_team + issue_owners}]")
                            xlsx2html(filename, sheet=ws, output='table.html')
                            body_table = open(os.path.abspath('table.html')).read()
                            mail_body = self.mailobj.speakup_mail(body_table, 'India')
                            flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=issue_owners, cc=ethics_team, bcc=BCC,
                                                         attachments=['sign.png', 'Picture1.png'])
                            if not flg:
                                self.logger.error("process_speakup(); Unable to send mail", extra=self.extra_dict_common)
                                return False
                    elif ws.__contains__('India'):
                        pass
                        """ethics_df = pd.read_excel(ethics_file, sheet_name='India', usecols=["ECO", "Region"])
                        region_dict = dict(ethics_df.values)
                        ethics_df.replace(r"^ +| +$", r"", regex=True, inplace=True)
                        india_df = xl.parse(ws)
                        india_df.replace(r"^ +| +$", r"", regex=True, inplace=True)
                        india_df['Region'] = india_df['Issue Owner'].map(region_dict)
                        india_df.to_excel("India.xlsx", 'input', index=False)
                        india_file = os.path.abspath("India.xlsx")
                        state = self.india_process_speakup(filename=india_file)
    
                        if not state:
                            return state
                        ind = pd.ExcelFile(india_file)
                        india_ws = ind.sheet_names
                        for iws in india_ws:
                            ind_owners = []
                            ind_df = ind.parse(iws)
                            emps = list(dict.fromkeys(ind_df['Issue Owner Employee Id'].tolist()))
                            for emp in emps:
                                ind_owners.append(self.file_obj.get_email_id(emp))
    
                            xlsx2html(india_file, sheet=iws, output='table.html')
                            body_table = open(os.path.abspath('table.html')).read()
                            mail_body = self.mailobj.speakup_mail(body_table, iws)
    
                            if iws.__contains__('India Ethics'):
                                ind_ethics_ids = list(np.setdiff1d(ind_ethics, ind_owners))
                                cc = ind_ethics_ids.extend(ethics_team)
                                flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=ind_owners, cc=cc,
                                                             attachments=['sign.png'])
                                if not flg:
                                    self.logger.error("process_speakup(); Unable to send mail",
                                                      extra=self.extra_dict_common)
                                    return False
                            elif iws.__contains__('India HR'):
    
                                ind_hr_ids = list(np.setdiff1d(ind_hr, ind_owners))
                                cc = ind_hr_ids.extend(ethics_team)
                                flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=ind_owners, cc=cc,
                                                             attachments=['sign.png'])
                                if not flg:
                                    self.logger.error("process_speakup(); Unable to send mail",
                                                      extra=self.extra_dict_common)
                                    return False
                            else:
                                flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=ind_owners, cc=ethics_team,
                                                             attachments=['sign.png'])
                                if not flg:
                                    self.logger.error("process_speakup(); Unable to send mail",
                                                      extra=self.extra_dict_common)
                                    return False"""
                    else:
                        self._delete_emp_id_(filename, ws)
                        val_df = pd.read_excel(filename, sheet_name=ws)
                        #encode_df = pd.read_excel(filename,sheet_name=ws,usecols=["Issue Owner"])
                        print(val_df)
                        if len(val_df) > 0:
                            ecos_list = list(ecos.get(ws, ''))
                            ecos_list.extend(ethics_team)
                            ecos_list = list(np.setdiff1d(ecos_list, issue_owners))
                            #df.to_html('table.html')
                            xlsx2html(filename, sheet=ws, output='table.html')
                            body_table = open(os.path.abspath('table.html')).read()
                            mail_body = self.mailobj.speakup_mail(body_table, ws)
                            flg = 1
                            flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=issue_owners, cc=ecos_list,
                                                         attachments=['sign.png', 'Picture1.png'])
                            self.status_file.write(str(mail_body[0] + " : " + "Success" + "\n"))
                            if not flg:
                                self.logger.error("process_speakup(); Unable to send mail", extra=self.extra_dict_common)
                                return False

                except Exception as err:
                    body_table = open(os.path.abspath('table.html')).read()
                    mail_body = self.mailobj.speakup_mail(body_table, ws)
                    self.status_file.write(str(mail_body[0] + " : " + "Failed" + "\n"))
                    self.logger.exception(f"process_speakup(); Exception while processing alerts; {err}",
                                          extra=self.extra_dict_common)
                    continue
        except Exception as err:
            self.logger.exception(f"process_speakup(); Exception while processing alerts; {err}",
                                  extra=self.extra_dict_common)
            return [False, err]

    def _delete_emp_id_(self, filename, ws, **krgs):
        try:
            book = openpyxl.load_workbook(filename)
            ws1 = book[ws]
            #ws1.delete_cols(9)
            ws1.delete_cols(6)
            #ws1.delete_cols(10)
            book.save(filename)
        except Exception as err:
            self.logger.exception(f"_delete_emp_id_(); Exception while deleting; {err}", extra=self.extra_dict_common)

    def india_process_speakup(self, filename, **krgs):

        try:
            ip_file = self.__split_data_by_team__(filename)
            if len(ip_file) > 1:
                return ip_file
            else:
                ip_file = ip_file[0]

            wb = load_workbook(filename)
            ws_list = wb.worksheets
            for ws in ws_list:
                self.set_border(ws, wb, ip_file)
                self.excel_font_styles(ip_file)
                flg = self.color_formatting(wb, ws, ip_file)
                if not flg:
                    return flg

            ip_file = self.set_header_color(ip_file)
            if len(ip_file) > 1:
                return ip_file
            else:
                ip_file = ip_file[0]

            flg = self.delete_region_column(ip_file)
            if not flg:
                return flg
            return True

        except Exception as err:
            self.logger.exception(f"india_process_speakup(); Exception while processing alerts; {err}",
                                  extra=self.extra_dict_common)
            return [False, err]

    def __split_data_by_team__(self, filename):

        try:
            countrywise_file = "india.xlsx"
            df = pd.read_excel(filename, sheet_name="input")
            df['Region'] = df['Region'].fillna('India')
            dict_of_country = dict(iter(df.groupby('Region')))
            writer = pd.ExcelWriter(countrywise_file, engine='openpyxl')
            for key, val in dict_of_country.items():
                val.to_excel(writer, sheet_name=key, header=True, index=False)

            writer.save()
            return [os.path.abspath(countrywise_file)]
        except Exception as err:
            self.logger.exception(f"__split_data_by_team__(); Failed to Process Data; {err}",
                                  extra=self.extra_dict_common)
            return [False, err]

    def add_india_teams_to_file(self, filename, ethics_file):

        try:
            ethics_df = pd.read_excel(ethics_file, sheet_name='India', usecols=["Region","Employee ID"])
            region_dict = dict(ethics_df.values)
            ethics_df.replace(r"^ +| +$", r"", regex=True, inplace=True)
            india_df = pd.read_excel(filename, sheet_name='India')
            india_df.replace(r"^ +| +$", r"", regex=True, inplace=True)
            #india_df['Region'] = india_df['Issue Owner'].map(region_dict)  # Issue Owner Employee Id
            india_df['Region'] = india_df['Issue Owner Employee Id'].map(region_dict)
            india_df.to_excel("India.xlsx", 'input', index=False)
            india_file = os.path.abspath("India.xlsx")
            return [india_file]
        except Exception as err:
            self.logger.exception(f"add_india_teams_to_file(); Exception: {err}", extra=self.extra_dict_common)
            return [False, err]

    def india_region_process_data(self, filename, ecos=None, ethics_team=None, ind_ethics=None, ind_hr=None,
                                  ethics_file=None, **krgs):
        try:
            #cc =ethics_team
            #ind = pd.ExcelFile(filename)
            ind  = openpyxl.load_workbook(filename)
            india_ws = ind.sheetnames
            for iws in india_ws:
                ind_owners = []
                #ind_df = ind.parse(iws)
                ind_df = pd.read_excel(filename,sheet_name=iws)
                emps = list(dict.fromkeys(ind_df['Issue Owner Employee Id'].tolist()))
                for emp in emps:
                    ind_owners.append(self.file_obj.get_email_id(emp))

                self._delete_emp_id_(filename, iws)
                xlsx2html(filename, sheet=iws, output='table.html')
                body_table = open(os.path.abspath('table.html')).read()
                mail_body = self.mailobj.speakup_mail(body_table, "India")

                if iws.__contains__('India Ethics'):
                    self.logger.info("***************INDIA ETHICS**************")
                    val_df = pd.read_excel(filename, sheet_name=iws)
                    if len(val_df)>0:
                        self.logger.info(f"india_region_process_data(); {iws}",
                                         extra=self.extra_dict_common)
                        ind_ethics_ids = list(np.setdiff1d(ind_ethics, ind_owners))
                        ind_ethics.extend(ethics_team)
                        ind_ethics = list(set(ind_ethics))

                        ind_ethics = list(np.setdiff1d(ind_ethics, ind_owners))

                        flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=ind_owners, cc=ind_ethics,
                                                     attachments=['sign.png', 'Picture1.png'])
                        self.status_file.write(str(mail_body[0] + " : " + "Success" + "\n"))
                        if not flg:
                            self.logger.error("process_speakup(); Unable to send mail",
                                              extra=self.extra_dict_common)
                            return False
                elif iws.__contains__('India HR'):
                    self.logger.info("***************INDIA HR**************")
                    val_df = pd.read_excel(filename, sheet_name=iws)
                    if len(val_df) > 0:
                        self.logger.info(f"india_region_process_data(); {iws}",
                                         extra=self.extra_dict_common)
                        ind_hr_ids = list(np.setdiff1d(ind_hr, ind_owners))

                        ind_hr.extend(ethics_team)
                        ind_hr = list(set(ind_hr))
                        ind_hr = list(np.setdiff1d(ind_hr, ind_owners))

                        flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=ind_owners, cc=ind_hr,
                                                     attachments=['sign.png', 'Picture1.png'])
                        self.status_file.write(str(mail_body[0] + " : " + "Success" + "\n"))
                        if not flg:
                            self.logger.error("process_speakup(); Unable to send mail",
                                              extra=self.extra_dict_common)
                            return False
                else:
                    self.logger.info("***************INDIA**************")
                    val_df = pd.read_excel(filename, sheet_name=iws)
                    if len(val_df) > 0:
                        self.logger.info(
                            f"ind_owners: {ind_owners} \n ethics_team: {ethics_team} \n",
                            extra=self.extra_dict_common)
                        self.logger.info(f"india_region_process_data(); {iws}",
                                         extra=self.extra_dict_common)

                        flg = self.mailobj.send_mail(mail_body[0], mail_body[1], to=ind_owners, cc=ethics_team,
                                                     attachments=['sign.png', 'Picture1.png'])
                        self.status_file.write(str(mail_body[0] + " : " + "Success" + "\n"))
                        if not flg:
                            self.logger.error("process_speakup(); Unable to send mail",
                                              extra=self.extra_dict_common)
                            return False
            return True
        except Exception as err:
            self.logger.exception(f"__india_region_process_data(); Exception: {err}", extra=self.extra_dict_common)
            return [False, err]

    def main(self):

        try:

            state = self.speakup_process()
            if state:
                extra_dict_status_success = {
                    "JobId": self.guid,
                    "logType": "User",
                    "exec_category": "Python_Script",
                    "Category": "Self Heal",
                    "businessCategory": "GI",
                    "useCaseName": "GI-Ethics",
                    "Status": "Successful",
                    "totalExecutionTimeInSeconds": time.time() - self.start_time,
                    "Complexity": "High",
                    "ManualEfforts": 100,
                }
                self.logger.info("------SUCCESS------",extra=extra_dict_status_success)
                exit(0)
            else:
                extra_dict_status_failed = {
                    "JobId": self.guid,
                    "logType": "User",
                    "exec_category": "Python_Script",
                    "Category": "Self Heal",
                    "businessCategory": "GI",
                    "useCaseName": "GI-Ethics",
                    "Status": "Failed",
                    "totalExecutionTimeInSeconds": time.time() - self.start_time,
                    "Complexity": "High",
                    "ManualEfforts": 0,
                }
                self.logger.exception(f"------FAILED-----", extra=extra_dict_status_failed)
                exit(1)

        except Exception as err:
            extra_dict_status_failed = {
                "JobId": self.guid,
                "logType": "User",
                "exec_category": "Python_Script",
                "Category": "Self Heal",
                "businessCategory": "GI",
                "useCaseName": "GI-Ethics",
                "Status": "Failed",
                "totalExecutionTimeInSeconds": time.time() - self.start_time,
                "Complexity": "High",
                "ManualEfforts": 0,
            }
            self.logger.exception(f"-----END----- Error: {err}", extra=extra_dict_status_failed)
            exit(1)


if __name__ == '__main__':
    obj = Main()
    obj.main()
