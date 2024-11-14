try:
    import uuid
    import os
    import glob
    import openpyxl
    import datetime
    import pandas as pd
    from openpyxl import load_workbook
    from logger_format import setup_logging
    from FileOperations import FileOperations
    from MailList import MailList
    from Config import FILEPATH
    from openpyxl.styles import Color, PatternFill, Font, Border, Side
    from xlsx2html import xlsx2html
except ImportError as error:
    print(f"Import Error..... Script Stopped.. {error}")
    exit(1)


class SpeakUp:

    def __init__(self, logFolder=None, logger=None, guid=None):

        if guid:
            self.guid = guid
        else:
            self.guid = uuid.uuid4()
        if logFolder:
            self.logger = logger
        else:
            self.logger = setup_logging(__file__)
        self.extra_dict_common = {
            "JobId": self.guid,
            "logType": "User",
            "exec_category": "Python_Script",
            "Category": "Self Heal",
            "businessCategory": "GI",
            "useCaseName": "GI-Ethics",
        }
        self.file_obj = FileOperations(__file__,self.logger, self.guid)
        self.mailobj = MailList(__file__,self.logger, self.guid)

    def filter_data_ownerwise(self, df):
        """
        Filter input dataframe by name Anne-Violaine AGAZZI-MONNIE and save it in different sheets.
        :param df:
        :return:
        """
        try:
            self.logger.info("Input data in DF received from shared folder", extra=self.extra_dict_common)
            ip_file = "input_data.xlsx"
            writer = pd.ExcelWriter(ip_file, engine='xlsxwriter')
            df[df['Issue Owner'] == "Anne-Violaine AGAZZI-MONNIE"].to_excel(writer, 'Anne-Violaine', index=False)
            df[df['Issue Owner'] != "Anne-Violaine AGAZZI-MONNIE"].to_excel(writer, 'input', index=False)
            writer.save()

            self.logger.info(f"Input data splited in two sheets and stored in {ip_file}", extra=self.extra_dict_common)
            return os.path.abspath(ip_file)
        except Exception as err:
            self.logger.exception(f"filter_data(); failed to filter out data; {err}", extra=self.extra_dict_common)
            return False

    def filter_countrywise_data(self, filename):

        try:
            self.logger.info("Filtering Data based on country", extra=self.extra_dict_common)
            countrywise_file = "input_data.xlsx"
            df = pd.read_excel(filename, sheet_name="input")
            violaine_df = pd.read_excel(filename, sheet_name="Anne-Violaine")
            dict_of_country = dict(iter(df.groupby('Country')))
            writer = pd.ExcelWriter(countrywise_file, engine='xlsxwriter')
            for key, val in dict_of_country.items():
                val.to_excel(writer, sheet_name=key, header=True, index=False)

            violaine_df.to_excel(writer, sheet_name='Anne-Violaine', header=True, index=False)
            writer.save()
            self.logger.info("Filtered Data based on country", extra=self.extra_dict_common)
            return os.path.abspath(countrywise_file)
        except Exception as err:
            self.logger.exception(" filter_countrywise_data(); Failed to Process Data", extra=self.extra_dict_common)
            return False

    def set_border(self, ws, wb, filename):

        border = Border(left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))
        #rows = ws.iter_rows()
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
            wb.save(filename)

    def set_header_color(self, fname):


        self.logger.info("set_header_color(); Setting color", extra= self.extra_dict_common)
        bg_clr = openpyxl.styles.colors.Color(rgb='000066CC')
        head_color = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=bg_clr)

        wb = load_workbook(fname)
        ws_list = wb.worksheets
        for ws in ws_list:
            for row in ws.iter_rows(max_row=1):
                for cel in row:
                    cel.font = Font(name='Calibri', color="00FFFFFF", size=10)
                    cel.fill = head_color
            wb.save(fname)

    def color_formatting(self, wb, ws, filename):
        """
        format excel cell based on rules like.

        Rule 1: Update Alert column text color to RED if created date is more than 45 days.
        Rule 2: Update D,E,F columns colours to Yellow if cell is blank.
        Rule 3: Update E column color to Yellow if cell containc XXX in value.
        :return:
        updated file name to share with business.
        """
        try:
            red_ft = Font(color="00FF0000")
            yellow_ft = openpyxl.styles.colors.Color(rgb='00FFF000')
            yellow_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=yellow_ft)
            for row in ws.iter_rows(min_row=2):
                try:
                    alert_cell = row[0]
                    date_cell = row[1]
                    sev_cell = row[3]
                    dept_cell = row[4]
                    issue_cell = row[5]

                    delta = datetime.datetime.today() - date_cell.value
                    if int(delta.days) > 30:
                        alert_cell.font = red_ft
                    if sev_cell.value is None:
                        sev_cell.fill = yellow_fill
                    if dept_cell.value is None or 'XXX' in str(dept_cell.value).upper():
                        dept_cell.fill = yellow_fill
                    if issue_cell.value is None:
                        issue_cell.fill = yellow_fill
                    wb.save(filename)

                except Exception as e:
                    self.logger.error(f"Failed to process row: {e}", extra=self.extra_dict_common)
            return True
        except Exception as err:
            self.logger.exception(f"Failed to process: {err}", extra=self.extra_dict_common)
            return err.__str__()

    def process(self):
        """
        Main process flow of Speakup Module is implemented here.
        :return:
        """
        try:
            self.logger.info("main(); Speakup Process Started;", extra=self.extra_dict_common)
            self.logger.info("main(); Getting File name from Shared path;", extra=self.extra_dict_common)
            response = self.file_obj.get_file_name(FILEPATH)
            if False in response:
                # Mail need to be send with error to business.
                resp = self.mailobj.file_missing_speakup(response[1])
                self.mailobj.send_mail(resp[0], resp[1])
                return False
            else:
                filename = response[1]
            self.logger.info(f"main(); reading data from file; {filename}", extra=self.extra_dict_common)
            fg,resp_df = self.file_obj.read_file(filename)
            print(resp_df)
            if False in fg:
                print('Mail need to be send with error to business')
                # Mail need to be send with error to business.
                resp = self.mailobj.file_missing_speakup(resp_df)
                self.mailobj.send_mail(resp[0], resp[1])
                return False

            self.logger.info(f"DF received with {resp_df.columns}", extra=self.extra_dict_common)
            ipfile = self.filter_data_ownerwise(resp_df)
            if ipfile is False:
                err = "Failed to filter data based on owners"
                resp = self.mailobj.file_missing_speakup(err)
                self.mailobj.send_mail(resp[0], resp[1])
                return False
            splitted_file = self.filter_countrywise_data(ipfile)
            #print(splitted_file)

            wb = load_workbook(splitted_file)
            ws_list = wb.worksheets
            for ws in ws_list:
                self.set_border(ws, wb, splitted_file)
                flag = self.color_formatting(wb, ws, splitted_file)
                print(ws, flag)
                if not flag:
                    resp = self.mailobj.file_missing_speakup(state)
                    self.mailobj.send_mail(resp[0], resp[1])
                    return False

            self.set_header_color(splitted_file)
            flag = self.speakup_process(splitted_file)
            if flag:
                self.logger.info("All mails sent to owner and also sent mail to business confirmation", extra=self.extra_dict_common)
            else:
                self.logger.error("Mail sending failed", extra=self.extra_dict_common)
                error = "Sending mail to issue owner and business confirmation failed."
                resp = self.mailobj.file_missing_speakup(error)
                self.mailobj.send_mail(resp[0], resp[1])
                return False

        except Exception as err:
            self.logger.exception(f"main(); Process Failed: {err}", extra=self.extra_dict_common)
            return False

    def speakup_process(self, filename):

        try:
            xl = pd.ExcelFile(filename)
            ws_list = xl.sheet_names
            emp_mails = []
            for ws in ws_list:
                df = xl.parse(ws)
                emps = list(dict.fromkeys(df['Issue Owner GGID'].tolist()))
                for emp in emps:
                    emp_mails.append(self.file_obj.get_email_id(emp))

                xlsx2html(filename, sheet=ws, output='table.html')
                body_table = open(os.path.abspath('table.html')).read()
                mail_body = self.mailobj.owner_report_mail(body_table)
                print(mail_body[0])
                flag = self.mailobj.send_mail(mail_body[0], mail_body[1], emp_mails)
                if not flag:
                    print("Unable to send mail to Issue owner")
                    self.logger.error("Unable to send mail", extra=self.extra_dict_common)
                    return False

            subject = "SpeakUp-FollowUp Report status- BOT process successful"
            mail_body = f"""
                        <p style="font-family: Verdana, Geneva, sans-serif;font-size:13px">
                            <p>
                                Hi Shantanu,<br><br>

                                SpeakUp follow Up reports have been successfully mailed to the identified recipients as of 
                                {datetime.datetime.now().strftime('%d-%m-%Y')}.
                            </p>
                            <p>
                                Regards,<br>
                                Autobot 
                            </p>

                            </p>
                        </p>
                        """
            flag = self.mailobj.send_mail(subject, mail_body, ['nanasaheb.yadav@capgemini.com'])
            if flag:
                self.logger.info("Procss success mail sent to business successfully", extra=self.extra_dict_common)
                return True
            else:
                print("failed to send mail to business")
                self.logger.error("Failed to send mail to business", extra=self.extra_dict_common)
                return False

        except Exception as e:
            self.logger.exception(f"Failed to process mail sending; {e}", extra=self.extra_dict_common)
            return False



if __name__ == '__main__':
    o = SpeakUp()
    state = o.process()
