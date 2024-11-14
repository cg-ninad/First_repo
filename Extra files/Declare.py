try:
    import uuid
    import os
    import glob
    import openpyxl
    import datetime
    import pandas as pd
    from openpyxl import load_workbook
    from logger_format import setup_logging
    from FileOperations import FileOperations
    from MailList import MailList
    from Config import FILEPATH, ETHICS_PATH, BUSINESS, CC, BCC, ARCHIEVE, SUPPORT
    from openpyxl.styles import Color, PatternFill, Font, Border, Side, Alignment
    from xlsx2html import xlsx2html

except ImportError as error:
    print(f"Import Error..... Script Stopped.. {error}")
    exit(1)


class Declare:

    def __init__(self, logFolder=None, logger=None, guid=None):

        if guid:
            self.guid = guid
        else:
            self.guid = uuid.uuid4()
        if logFolder:
            self.logger = logger
        else:
            self.logger = setup_logging(__file__)
        self.extra_dict_common = {
            "JobId": self.guid,
            "logType": "User",
            "exec_category": "Python_Script",
            "Category": "Self Heal",
            "businessCategory": "GI",
            "useCaseName": "GI-Ethics",
        }
        self.file_obj = FileOperations(__file__, self.logger, self.guid)
        self.mailobj = MailList(__file__,self.logger, self.guid)

    def declare_process(self, **krgs):

        """
        Declare Process start from here
        :param krgs:
        :return:
        """
        try:

            self.logger.info("declare_process(); Declare Process Started;", extra=self.extra_dict_common)
            self.logger.info("declare_process(); Getting File name from Shared path;", extra=self.extra_dict_common)

            ethics_file = self.file_obj.get_file_name(ETHICS_PATH)
            print("1. ", ethics_file)
            if False in ethics_file:
                self.logger.error(f"declare_process(); File not found: {ethics_file}", extra=self.extra_dict_common)
                resp = self.mailobj.file_missing_declare(ethics_file[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                ethics_file = ethics_file[1]
                print("2. ", ethics_file)

            response = self.file_obj.get_file_name(FILEPATH)
            print("3. ", response)
            if False in response:
                self.logger.error(f"declare_process(); File not found: {response}", extra=self.extra_dict_common)
                resp = self.mailobj.file_missing_declare(response[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                filename = response[1]
                self.logger.info(f"declare_process(); File to process: {filename}", extra=self.extra_dict_common)
                if 'declare' not in str(filename).lower():
                    msg = "Declare file not found in folder, Invalid file is uploaded. Please upload Declare " \
                          f"alerts file including <mark>'declare'</mark> in filename.<br> file uploaded: {filename}"
                    resp = self.mailobj.file_missing_declare(msg)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False

            self.logger.info(f"declare_process(); reading data from file; {filename}", extra=self.extra_dict_common)
            resp_df = self.file_obj.read_file(filename)

            if False in resp_df:
                # Mail need to be send with error to business.
                resp = self.mailobj.file_missing_declare(resp_df[1])
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False

            self.logger.info(f"Columns: {resp_df.columns.to_list()}", extra=self.extra_dict_common)
            df = self.add_region_in_df_basedon_code(resp_df, ethics_file)
            print("4. ", df.columns.tolist())
            if False in df:
                resp = self.mailobj.file_missing_declare(df)
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False

            self.logger.info(f"DF received with {resp_df.columns}", extra=self.extra_dict_common)
            ip_file = self.split_alerts_by_region(df)
            print("5. ", ip_file)
            if False in ip_file:
                resp = self.mailobj.file_missing_declare(df)
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            else:
                ip_file = ip_file[0]

            wb = load_workbook(ip_file)
            ws_list = wb.worksheets
            for ws in ws_list:
                self.set_border(ws, wb, ip_file)
                self.excel_font_styles(ip_file)
                flag = self.color_formatting(wb, ws, ip_file)
                #print(ws, flag)
                if not flag:
                    resp = self.mailobj.file_missing_declare(flag)
                    self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                    return False
            self.set_header_color(ip_file)

            flg = self.delete_region_column(ip_file)
            if not flg:
                resp = self.mailobj.file_missing_declare(flg)
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            owner_df = pd.read_excel(ethics_file, sheet_name='ECO', usecols=["Region", "Email Address"])
            ethic_df = pd.read_excel(ethics_file, sheet_name='Ethics Team', usecols=["Email Address"])
            cc_list = ethic_df['Email Address'].to_list()
            owner_dict = {k: g["Email Address"].tolist() for k,g in owner_df.groupby("Region")}

            flg= self.process_alerts(ip_file, owner_dict, cc_list)

            if False in flg:
                self.logger.error("declare_process(); Mail not sent to flg;", extra=self.extra_dict_common)
                resp = self.mailobj.file_missing_declare(flg)
                self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
                return False
            subject, mail_body = self.mailobj.declare_success()
            status = self.mailobj.send_mail(subject, mail_body, BUSINESS, cc=CC, bcc=BCC)
            if status:
                self.logger.info("Mail send to business success", extra=self.extra_dict_common)
                resp = self.file_obj.move_files_archive(filename, ARCHIEVE)
                self.logger.info(f"File Archived status: {resp}", extra=self.extra_dict_common)
                return True
            else:
                self.logger.error("Failed to send mail to business success", extra=self.extra_dict_common)
                return False
        except Exception as err:
            self.logger.exception(f"declare_process(); Exception: {err}", extra=self.extra_dict_common)
            resp = self.mailobj.file_missing_declare(err)
            self.mailobj.send_mail(resp[0], resp[1], to=CC, cc=BUSINESS, bcc=BCC)
            return False

    def add_region_in_df_basedon_code(self, declare_df=None, ethics_file=None, **krgs):

        try:
            ethics_df = pd.read_excel(ethics_file, usecols=["Country Code", "Region"])
            region_dict = dict(ethics_df.values)
            declare_df['Region'] = declare_df['Country'].map(region_dict)
            return declare_df
        except Exception as excpt:
            self.logger.exception(f"add_region_in_df_basedon(); Exception: {excpt}", extra=self.extra_dict_common)
            return [False, excpt.__str__()]

    def split_alerts_by_region(self, df, **krgs):

        try:
            region_file = "DeclareInput.xlsx"
            dict_of_country = dict(iter(df.groupby('Region')))
            writer = pd.ExcelWriter(region_file, engine='xlsxwriter')
            for key, val in dict_of_country.items():
                val.to_excel(writer, sheet_name=key, header=True, index=False)
            writer.save()
            self.logger.info("Filtered Data based on region", extra=self.extra_dict_common)
            return [os.path.abspath(region_file)]
        except Exception as err:
            self.logger.exception(f"split_alerts_by_region(); Exception: {err}", extra=self.extra_dict_common)
            return [False, err]

    def set_border(self, ws, wb, filename):

        border = Border(left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))
        text_ft = Font(name='Verdana', size=12)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.font = text_ft
                cell.alignment = Alignment(horizontal='center', vertical='center')
            wb.save(filename)

    def color_formatting(self, wb, ws, filename):
        """
        format excel cell based on rules like.

        Rule 1: Update Alert column text color to RED if created date is more than 30 days.
        :return:
        updated file name to share with business.
        """
        try:
            red_ft = Font(color="00FF0000", name='Verdana', size=12)
            for row in ws.iter_rows(min_row=2):
                try:
                    alert_cell = row[0]
                    date_cell = row[1]
                    delta = datetime.datetime.today() - date_cell.value
                    if int(delta.days) > 30:
                        alert_cell.font = red_ft
                    wb.save(filename)
                except Exception as e:
                    self.logger.error(f"color_formatting(); failed to process row: {e}", extra=self.extra_dict_common)
            return True
        except Exception as err:
            self.logger.exception(f"Failed to process: {err}", extra=self.extra_dict_common)
            return err.__str__()

    def set_header_color(self, fname):

        self.logger.info("set_header_color(); Setting color", extra= self.extra_dict_common)
        bg_clr = openpyxl.styles.colors.Color(rgb='000066CC')
        head_color = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=bg_clr)

        border = Border(left=Side(border_style='medium'),
                        right=Side(border_style='medium'),
                        top=Side(border_style='medium'),
                        bottom=Side(border_style='medium'))

        wb = load_workbook(fname)
        ws_list = wb.worksheets
        for ws in ws_list:
            for row in ws.iter_rows(max_row=1):
                for cel in row:
                    cel.font = Font(name='Verdana', color="00FFFFFF", size=12, bold=True)
                    cel.fill = head_color
                    #cel.border = border
            wb.save(fname)

    def process_alerts(self,filename, to_owners, cc_ethics):
        """
        :param filename:
        :return:
        """
        try:
            xl = pd.ExcelFile(filename)
            ws_list = xl.sheet_names

            for ws in ws_list:
                xlsx2html(filename, sheet=ws, output='declare.html')
                body_table = open(os.path.abspath('declare.html')).read()
                print("sheet name:", ws)
                mail_body = self.mailobj.declare_mail(body_table, ws)
                owners = to_owners.get(ws,'')
                owners = list(set(owners))
                cc_ethics = list(set(cc_ethics))
                print("alert owners: ", owners)
                print("Ethics Team: ", cc_ethics)
                flag = self.mailobj.send_mail(mail_body[0], mail_body[1], owners, cc_ethics, attachments=['sign.png'])
                if not flag:
                    print("Unable to send mail to Issue owner")
                    self.logger.error(f"Failed to send mail to {owners} for region {region}", extra=self.extra_dict_common)
                    msg = f"""Failed to send mail to '{', '.join(owners)}' of region {ws}. 
                            Please do necessary actions for respective region. 
                            """
                    resp = self.mailobj.file_missing_declare(msg)
                    self.mailobj.send_mail(resp[0], resp[1], to=BUSINESS, cc=CC, bcc=BCC)
                    #return [False, f"Failed to send mail to {owners} of region {ws}"]
                self.logger.info(f"process_alerts(); mail send to {ws} owners", extra=self.extra_dict_common)

            return [True]
        except Exception as err:
            self.logger.exception(f"process_alerts(); Exception: {err}", extra=self.extra_dict_common)
            return [False, err]

    def delete_region_column(self, filename):

        try:
            book = openpyxl.load_workbook(filename)
            ws_list = book.worksheets
            # delete column from existing sheet
            for ws in ws_list:
                ws.delete_cols(6)
                book.save(filename)
            return True
        except Exception as err:
            self.logger.exception(f"Delete Col Failed: {err}", extra=self.extra_dict_common)
            return [False, err]

    def excel_font_styles(self, filename):

        try:
            font = Font(name='Calibri', size=11, bold=False, italic=False, vertAlign=None, underline='none',
                        strike=False)
            border = Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
            wb = load_workbook(filename)
            ws_list = wb.worksheets
            for ws in ws_list:
                for row in ws.iter_rows(min_row=2):
                    for cel in row:
                        cel.font = font
                        cel.border = border
                wb.save(filename)
            return True
        except Exception as err:
            self.logger.exception(f"excel_font_styles(); Exception: {err}", extra=self.extra_dict_common)
            return False, err


o = Declare()
s = o.declare_process()
print(s)
